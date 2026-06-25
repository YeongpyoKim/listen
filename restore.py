import json
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent
JSON_PATH = ROOT / "site" / "data" / "stores.json"

def main():
    if not JSON_PATH.exists():
        print(f"Error: {JSON_PATH} not found.")
        return

    with open(JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    stores = data.get("stores", [])
    
    # 1. Generate site_content.py
    content_dict = {}
    for s in stores:
        name = s["name"]
        content_dict[name] = {
            "emoji": s.get("emoji", "🍽️"),
            "accent": s.get("accent", "#8a6d3b"),
            "tagline": s.get("tagline", ""),
            "story": s.get("story", ""),
            "sunday": s.get("sunday", "check"),
            "sunday_note": s.get("sunday_note", ""),
            "address": s.get("address", ""),
            "phone": s.get("phone", ""),
        }
    
    content_py_path = ROOT / "site_content.py"
    with open(content_py_path, "w", encoding="utf-8") as f:
        f.write("# -*- coding: utf-8 -*-\n")
        f.write("CONTENT = ")
        f.write(json.dumps(content_dict, ensure_ascii=False, indent=4))
        f.write("\n")
    print(f"[+] Reconstructed {content_py_path}")

    # 2. Generate 상점 리스트.xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "상점리스트"
    
    # Add some dummy headers
    ws["A3"] = "번호"
    ws["B3"] = "상점명"
    ws["C3"] = "구분"
    ws["D3"] = "대표자"
    ws["E3"] = "대표메뉴"
    ws["F3"] = "가격대"
    ws["G3"] = "영업시간"
    ws["H3"] = "좌석형태"
    ws["I3"] = "아기의자"
    ws["J3"] = "주차"
    ws["K3"] = "역사"
    ws["L3"] = "팁"
    ws["M3"] = "매력"
    
    # Write store records
    for idx, s in enumerate(stores, start=1):
        row_num = 3 + idx
        ws.cell(row=row_num, column=1, value=idx)
        
        name_cell = ws.cell(row=row_num, column=2, value=s["name"])
        if s.get("naver_url"):
            name_cell.hyperlink = s["naver_url"]
            name_cell.style = "Hyperlink"
            
        ws.cell(row=row_num, column=3, value=s.get("type", ""))
        ws.cell(row=row_num, column=4, value="") # manager
        ws.cell(row=row_num, column=5, value=s.get("signature", ""))
        ws.cell(row=row_num, column=6, value=s.get("price", ""))
        ws.cell(row=row_num, column=7, value=s.get("hours", ""))
        ws.cell(row=row_num, column=8, value=s.get("seats", ""))
        ws.cell(row=row_num, column=9, value=s.get("baby_chair", ""))
        ws.cell(row=row_num, column=10, value=s.get("parking", ""))
        ws.cell(row=row_num, column=11, value=s.get("history", ""))
        ws.cell(row=row_num, column=12, value=s.get("tip", ""))
        ws.cell(row=row_num, column=13, value=s.get("charm", ""))

    xlsx_path = ROOT / "상점 리스트.xlsx"
    wb.save(xlsx_path)
    print(f"[+] Reconstructed {xlsx_path}")

if __name__ == "__main__":
    main()
